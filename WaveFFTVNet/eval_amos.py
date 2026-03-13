import os
import re
import json
import math
import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np
import torch
from torch.amp import autocast
from monai.data import CacheDataset, DataLoader, decollate_batch
from monai.inferers import sliding_window_inference
from monai.transforms import (
    AsDiscrete,
    Compose,
    CropForegroundd,
    EnsureChannelFirstd,
    LoadImaged,
    Orientationd,
    ScaleIntensityRanged,
    Spacingd,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from networks.model import VNet


CLASS_LABELS = {
    0: "background",
    1: "spleen",
    2: "rkid",
    3: "lkid",
    4: "gall",
    5: "eso",
    6: "liver",
    7: "sto",
    8: "aorta",
    9: "IVC",
    10: "pancreas",
    11: "rad",
    12: "lad",
    13: "duodenum",
    14: "bladder",
    15: "prostate_or_uterus",
}


DISPLAY_NAME_MAP = {
    1: "Spleen",
    2: "R.Kd",
    3: "L.Kd",
    4: "GB",
    5: "Eso.",
    6: "Liver",
    7: "Stom.",
    8: "Aorta",
    9: "IVC",
    10: "Panc.",
    11: "RAG",
    12: "LAG",
    13: "Duo.",
    14: "Blad.",
    15: "Pros.",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Evaluate AMOS/BTCV-style test set with MONAI sliding window inference "
            "and export summary to Excel."
        )
    )
    parser.add_argument("--data_dir", type=str, required=True, help="数据根目录，包含 imagesTr/labelsTr")
    parser.add_argument(
        "--test_list_json",
        type=str,
        required=True,
        help="训练脚本导出的 test_files_list.json",
    )
    parser.add_argument(
        "--output_xlsx",
        type=str,
        default="",
        help="输出 Excel 路径；默认写到 test_list 同目录下 eval_summary.xlsx",
    )
    parser.add_argument(
        "--method_name",
        type=str,
        default="",
        help="Excel 中 Method 列显示的名字；默认自动从 checkpoint 推断",
    )

    parser.add_argument(
        "--ckpt",
        type=str,
        nargs="*",
        default=[],
        help="一个或多个模型权重；多个权重时做 logits 平均集成",
    )
    parser.add_argument(
        "--topk_json",
        type=str,
        default="",
        help="可直接读取训练导出的 topk_checkpoints.json",
    )

    parser.add_argument("--num_classes", type=int, default=16)
    parser.add_argument("--pixdim", type=float, nargs=3, default=[1.5, 1.5, 2.0])
    parser.add_argument("--roi_size", type=int, nargs=3, default=[96, 96, 96])
    parser.add_argument("--sw_batch_size", type=int, default=2)
    parser.add_argument("--sw_overlap", type=float, default=0.5)

    parser.add_argument("--cache_rate", type=float, default=1.0)
    parser.add_argument("--num_workers", type=int, default=4)
    parser.add_argument("--amp", action="store_true", help="仅在 CUDA 上启用混合精度")
    parser.add_argument(
        "--device",
        type=str,
        default="cuda",
        help="cuda / cpu；若指定 cuda 但不可用，会自动退回 cpu",
    )
    return parser.parse_args()


def safe_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    x = float(x)
    if math.isnan(x) or math.isinf(x):
        return None
    return x


def extract_case_ids(text: str) -> List[str]:
    text = str(text)
    ids4 = re.findall(r"(?<!\d)(\d{4})(?!\d)", text)
    if ids4:
        return ids4
    out: List[str] = []
    for g in re.findall(r"\d+", text):
        if len(g) <= 4:
            out.append(g.zfill(4))
    return out


def replace_dir_token(path_str: str, old_token: str, new_token: str) -> str:
    pattern = rf"([/\\\\]){re.escape(old_token)}([/\\\\])"
    return re.sub(pattern, rf"\\1{new_token}\\2", path_str, count=1)


def resolve_path(candidate: str, roots: Sequence[Path]) -> Optional[Path]:
    if not candidate:
        return None

    raw = Path(candidate)
    to_try: List[Path] = []
    if raw.is_absolute():
        to_try.append(raw)
    else:
        to_try.append(raw)
        for r in roots:
            to_try.append(r / raw)

    # imagesTs/labelsTs -> imagesTr/labelsTr 回退
    replaced: List[Path] = []
    for p in list(to_try):
        s = str(p)
        s = replace_dir_token(s, "imagesTs", "imagesTr")
        s = replace_dir_token(s, "labelsTs", "labelsTr")
        replaced.append(Path(s))
    to_try.extend(replaced)

    seen = set()
    unique_try: List[Path] = []
    for p in to_try:
        rp = str(p)
        if rp not in seen:
            seen.add(rp)
            unique_try.append(p)

    for p in unique_try:
        if p.exists():
            return p.resolve()
    return None


def search_in_tr(case_id: str, folder: Path) -> Optional[Path]:
    patterns = [
        f"*{case_id}*.nii.gz",
        f"*{case_id}*.nii",
        f"*{case_id}*.mha",
        f"*{case_id}*.nrrd",
    ]
    matches: List[Path] = []
    for pat in patterns:
        matches.extend(sorted(folder.rglob(pat)))
    matches = [p.resolve() for p in matches if p.is_file()]

    uniq: List[Path] = []
    seen = set()
    for p in matches:
        sp = str(p)
        if sp not in seen:
            seen.add(sp)
            uniq.append(p)

    if len(uniq) == 1:
        return uniq[0]
    if len(uniq) > 1:
        # 优先 amos_xxxx 命名
        exact = [p for p in uniq if f"_{case_id}." in p.name or p.stem.endswith(case_id)]
        if len(exact) == 1:
            return exact[0]
        raise RuntimeError(
            f"在 {folder} 中找到多个匹配 case_id={case_id} 的文件：\n" + "\n".join(str(p) for p in uniq)
        )
    return None


def resolve_test_pairs(test_list_json: str, data_dir: str) -> List[Dict[str, Any]]:
    test_list_path = Path(test_list_json).resolve()
    data_root = Path(data_dir).resolve()
    list_root = test_list_path.parent
    images_tr = data_root / "imagesTr"
    labels_tr = data_root / "labelsTr"

    with open(test_list_path, "r", encoding="utf-8") as f:
        entries = json.load(f)

    if not isinstance(entries, list) or len(entries) == 0:
        raise ValueError(f"test_list_json 非法或为空: {test_list_json}")

    roots = [list_root, data_root]
    resolved: List[Dict[str, Any]] = []
    for i, d in enumerate(entries):
        image_raw = str(d.get("image", "") or "")
        label_raw = str(d.get("label", "") or "")

        case_ids = extract_case_ids(os.path.basename(image_raw))
        if not case_ids:
            case_ids = extract_case_ids(os.path.basename(label_raw))
        if not case_ids:
            case_ids = extract_case_ids(json.dumps(d, ensure_ascii=False))
        if not case_ids:
            raise RuntimeError(f"无法从第 {i} 个样本中解析 case id: {d}")
        case_id = case_ids[0]

        image_path = resolve_path(image_raw, roots)
        label_path = resolve_path(label_raw, roots)

        # 强制优先 imagesTr/labelsTr
        if image_path is None or "imagesTr" not in image_path.parts:
            found = search_in_tr(case_id, images_tr)
            if found is not None:
                image_path = found
        if label_path is None or "labelsTr" not in label_path.parts:
            found = search_in_tr(case_id, labels_tr)
            if found is not None:
                label_path = found

        if image_path is None:
            raise FileNotFoundError(
                f"无法解析 image 文件: case_id={case_id}, 原始条目={image_raw}, 搜索目录={images_tr}"
            )
        if label_path is None:
            raise FileNotFoundError(
                f"无法解析 label 文件: case_id={case_id}, 原始条目={label_raw}, 搜索目录={labels_tr}"
            )

        resolved.append(
            {
                "image": str(image_path),
                "label": str(label_path),
                "case_id": case_id,
                "orig_image": image_raw,
                "orig_label": label_raw,
            }
        )
    return resolved


def get_eval_transforms(pixdim: Tuple[float, float, float]) -> Compose:
    return Compose(
        [
            LoadImaged(keys=["image", "label"]),
            EnsureChannelFirstd(keys=["image", "label"]),
            Orientationd(
                keys=["image", "label"],
                axcodes="RAS",
                labels=(("L", "R"), ("P", "A"), ("I", "S")),
            ),
            Spacingd(keys=["image", "label"], pixdim=pixdim, mode=("bilinear", "nearest")),
            ScaleIntensityRanged(
                keys=["image"],
                a_min=-125,
                a_max=275,
                b_min=0.0,
                b_max=1.0,
                clip=True,
            ),
            CropForegroundd(keys=["image", "label"], source_key="image", allow_smaller=True),
        ]
    )


def dice_per_class_onehot(
    y_pred_1hot: torch.Tensor,
    y_true_1hot: torch.Tensor,
    eps: float = 1e-8,
    ignore_empty: bool = True,
) -> torch.Tensor:
    if y_pred_1hot.dim() == 5:
        y_pred_1hot = y_pred_1hot[0]
    if y_true_1hot.dim() == 5:
        y_true_1hot = y_true_1hot[0]

    y_pred = y_pred_1hot.float()
    y_true = y_true_1hot.float()
    dims = tuple(range(1, y_pred.dim()))

    inter = (y_pred * y_true).sum(dim=dims)
    pred_sum = y_pred.sum(dim=dims)
    true_sum = y_true.sum(dim=dims)
    denom = pred_sum + true_sum
    dice = (2.0 * inter) / (denom + eps)

    if ignore_empty:
        nan = torch.tensor(float("nan"), device=dice.device, dtype=dice.dtype)
        dice = torch.where(true_sum > 0, dice, nan)
    else:
        dice = torch.where(denom > 0, dice, torch.ones_like(dice))
    return dice


def load_state_dict_flex(ckpt_path: str, device: torch.device) -> Dict[str, torch.Tensor]:
    obj = torch.load(ckpt_path, map_location=device)
    if isinstance(obj, dict) and "model" in obj and isinstance(obj["model"], dict):
        return obj["model"]
    if isinstance(obj, dict):
        return obj
    raise RuntimeError(f"不支持的 checkpoint 格式: {ckpt_path}")


def resolve_ckpt_paths(args: argparse.Namespace) -> List[str]:
    ckpt_paths: List[str] = []
    if args.ckpt:
        ckpt_paths.extend(args.ckpt)
    elif args.topk_json:
        topk_json = Path(args.topk_json)
        with open(topk_json, "r", encoding="utf-8") as f:
            rows = json.load(f)
        for row in rows:
            p = row.get("path", "")
            if p:
                ckpt_paths.append(p)
    else:
        base = Path(args.test_list_json).resolve().parent
        auto_topk = base / "topk_checkpoints.json"
        auto_best = base / "best.pt"
        if auto_topk.is_file():
            with open(auto_topk, "r", encoding="utf-8") as f:
                rows = json.load(f)
            for row in rows:
                p = row.get("path", "")
                if p:
                    ckpt_paths.append(p)
        elif auto_best.is_file():
            ckpt_paths.append(str(auto_best))

    ckpt_paths = [str(Path(p).resolve()) for p in ckpt_paths if Path(p).is_file()]
    if not ckpt_paths:
        raise FileNotFoundError(
            "未找到可用 checkpoint。请显式传入 --ckpt 或 --topk_json，"
            "或者将 eval.py 放在训练输出目录并保证同目录存在 best.pt / topk_checkpoints.json。"
        )
    return ckpt_paths


def build_models(ckpt_paths: Sequence[str], device: torch.device, num_classes: int) -> List[torch.nn.Module]:
    models: List[torch.nn.Module] = []
    for p in ckpt_paths:
        model = VNet(n_channels=1, n_classes=num_classes).to(device)
        state_dict = load_state_dict_flex(p, device)
        model.load_state_dict(state_dict, strict=True)
        model.eval()
        models.append(model)
    return models


@torch.no_grad()
def evaluate_cases(
    models: Sequence[torch.nn.Module],
    loader: DataLoader,
    device: torch.device,
    roi_size: Tuple[int, int, int],
    sw_batch_size: int,
    sw_overlap: float,
    num_classes: int,
    use_amp: bool,
) -> List[Dict[str, Any]]:
    post_label = AsDiscrete(to_onehot=num_classes)
    post_pred = AsDiscrete(argmax=True, to_onehot=num_classes)

    case_rows: List[Dict[str, Any]] = []
    for batch in loader:
        inputs = batch["image"].to(device)
        labels = batch["label"].to(device)
        case_id = str(batch["case_id"][0])
        image_path = str(batch["image_path"][0])
        label_path = str(batch["label_path"][0])

        logits_sum = None
        with autocast(device_type=device.type, enabled=use_amp):
            for model in models:
                logits = sliding_window_inference(
                    inputs,
                    roi_size=roi_size,
                    sw_batch_size=sw_batch_size,
                    predictor=model,
                    overlap=sw_overlap,
                    mode="gaussian",
                )
                logits_sum = logits if logits_sum is None else (logits_sum + logits)

        logits = logits_sum / float(len(models))

        labels_list = decollate_batch(labels)
        logits_list = decollate_batch(logits)
        labels_1hot = post_label(labels_list[0])
        preds_1hot = post_pred(logits_list[0])

        dice_c = dice_per_class_onehot(preds_1hot, labels_1hot, ignore_empty=True).detach().cpu().float().numpy()
        fg_mean = float(np.nanmean(dice_c[1:])) if num_classes > 1 else float(np.nanmean(dice_c))

        row = {
            "case_id": case_id,
            "image": image_path,
            "label": label_path,
            "dice_per_class_incl_bg": dice_c.tolist(),
            "dice_mean_fg": fg_mean,
        }
        case_rows.append(row)
        print(f"[CASE] {case_id} | fg_dice={fg_mean:.6f}")

    return case_rows


def aggregate_all_cases(
    case_rows: List[Dict[str, Any]],
    num_classes: int,
) -> Tuple[np.ndarray, float]:
    if len(case_rows) == 0:
        raise ValueError("没有可评估样本")

    arr = np.full((len(case_rows), num_classes), np.nan, dtype=np.float32)
    for i, row in enumerate(case_rows):
        arr[i] = np.asarray(row["dice_per_class_incl_bg"], dtype=np.float32)

    per_class_mean = np.nanmean(arr, axis=0)
    mean_fg = float(np.nanmean(per_class_mean[1:])) if num_classes > 1 else float(np.nanmean(per_class_mean))
    return per_class_mean, mean_fg


def infer_method_name(method_name: str, ckpt_paths: Sequence[str]) -> str:
    if method_name:
        return method_name
    if len(ckpt_paths) == 1:
        return Path(ckpt_paths[0]).stem
    return f"Ensemble({len(ckpt_paths)})"


def beautify_worksheet(ws) -> None:
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000"
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for c in col_cells:
            v = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)

    ws.freeze_panes = "A2"


def save_excel(
    output_xlsx: str,
    method_name: str,
    per_class_mean: np.ndarray,
    mean_fg: float,
    case_rows: List[Dict[str, Any]],
    ckpt_paths: Sequence[str],
    args: argparse.Namespace,
) -> None:
    wb = Workbook()

    # 1) summary sheet
    ws = wb.active
    ws.title = "summary"
    organ_headers = [DISPLAY_NAME_MAP[i] for i in range(1, args.num_classes)]
    summary_headers = ["Method"] + organ_headers + ["Average"]
    ws.append(summary_headers)

    summary_row = [method_name]
    for c in range(1, args.num_classes):
        summary_row.append(safe_float(per_class_mean[c]))
    summary_row.append(safe_float(mean_fg))
    ws.append(summary_row)
    beautify_worksheet(ws)

    # 2) case_details sheet
    ws_case = wb.create_sheet("case_details")
    case_headers = ["CaseID", "Average"] + organ_headers + ["Image", "Label"]
    ws_case.append(case_headers)

    for row in case_rows:
        dice_vec = np.asarray(row["dice_per_class_incl_bg"], dtype=np.float32)
        excel_row = [
            row["case_id"],
            safe_float(row["dice_mean_fg"]),
        ]
        for c in range(1, args.num_classes):
            excel_row.append(safe_float(dice_vec[c]))
        excel_row.extend([row["image"], row["label"]])
        ws_case.append(excel_row)
    beautify_worksheet(ws_case)

    # 3) meta sheet
    ws_meta = wb.create_sheet("meta")
    ws_meta.append(["key", "value"])
    meta_rows = [
        ("num_test_cases_total", len(case_rows)),
        ("final_mean_fg", safe_float(mean_fg)),
        ("data_dir", str(Path(args.data_dir).resolve())),
        ("test_list_json", str(Path(args.test_list_json).resolve())),
        ("roi_size", json.dumps(list(args.roi_size))),
        ("pixdim", json.dumps(list(args.pixdim))),
        ("sw_batch_size", args.sw_batch_size),
        ("sw_overlap", args.sw_overlap),
        ("num_classes", args.num_classes),
        ("device", args.device),
        ("amp", bool(args.amp)),
        ("ckpt_paths", "\n".join([str(Path(p).resolve()) for p in ckpt_paths])),
    ]
    for k, v in meta_rows:
        ws_meta.append([k, v])
    beautify_worksheet(ws_meta)

    out_path = Path(output_xlsx).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"[SAVE] Excel -> {out_path}")


def main() -> None:
    args = parse_args()

    if args.device == "cuda" and not torch.cuda.is_available():
        print("[WARN] CUDA 不可用，自动切换到 CPU")
        args.device = "cpu"

    device = torch.device(args.device)
    use_amp = bool(args.amp and device.type == "cuda")

    if not args.output_xlsx:
        args.output_xlsx = str(Path(args.test_list_json).resolve().parent / "eval_summary.xlsx")

    ckpt_paths = resolve_ckpt_paths(args)
    method_name = infer_method_name(args.method_name, ckpt_paths)

    print("=" * 80)
    print(f"Method         : {method_name}")
    print(f"Checkpoint(s)  : {len(ckpt_paths)}")
    for i, p in enumerate(ckpt_paths):
        print(f"  [{i}] {p}")
    print(f"Device         : {device}")
    print(f"AMP            : {use_amp}")
    print(f"ROI size       : {tuple(args.roi_size)}")
    print(f"Pixdim         : {tuple(args.pixdim)}")
    print(f"SW batch size  : {args.sw_batch_size}")
    print(f"SW overlap     : {args.sw_overlap}")
    print("=" * 80)

    test_files = resolve_test_pairs(args.test_list_json, args.data_dir)
    print(f"[INFO] resolved test cases: {len(test_files)}")

    dataset = CacheDataset(
        data=[
            {
                "image": row["image"],
                "label": row["label"],
                "case_id": row["case_id"],
                "image_path": row["image"],
                "label_path": row["label"],
            }
            for row in test_files
        ],
        transform=get_eval_transforms(tuple(args.pixdim)),
        cache_num=len(test_files),
        cache_rate=float(args.cache_rate),
        num_workers=int(args.num_workers),
    )

    loader = DataLoader(
        dataset,
        batch_size=1,
        shuffle=False,
        num_workers=int(args.num_workers),
        pin_memory=(device.type == "cuda"),
        persistent_workers=(int(args.num_workers) > 0),
    )

    models = build_models(ckpt_paths, device, args.num_classes)

    case_rows = evaluate_cases(
        models=models,
        loader=loader,
        device=device,
        roi_size=tuple(args.roi_size),
        sw_batch_size=int(args.sw_batch_size),
        sw_overlap=float(args.sw_overlap),
        num_classes=int(args.num_classes),
        use_amp=use_amp,
    )

    per_class_mean, mean_fg = aggregate_all_cases(
        case_rows=case_rows,
        num_classes=int(args.num_classes),
    )

    print("\n[RESULT] final summary on all test cases:")
    for c in range(1, args.num_classes):
        val = per_class_mean[c]
        print(f"  {DISPLAY_NAME_MAP[c]:>6s} : {float(val):.6f}")
    print(f"  {'Average':>6s} : {mean_fg:.6f}")

    save_excel(
        output_xlsx=args.output_xlsx,
        method_name=method_name,
        per_class_mean=per_class_mean,
        mean_fg=mean_fg,
        case_rows=case_rows,
        ckpt_paths=ckpt_paths,
        args=args,
    )


if __name__ == "__main__":
    main()